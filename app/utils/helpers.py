import sqlite3
import datetime
import os
import re
import pandas as pd
import hashlib
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Font

# Configuração do banco de dados
db_path = os.path.abspath(os.path.join(os.path.dirname(__file__), '../../data/Micas_Daminhas.db'))

def get_db_connection():
    if not os.path.exists(db_path):
        with open(db_path, 'w'):
            pass
    conn = sqlite3.connect(db_path, check_same_thread=False)
    return conn

def hash_password(password):
    return hashlib.sha256(password.encode()).hexdigest()

def initialize_database():
    # Criar tabela se não existir
    with get_db_connection() as conn:
        c = conn.cursor()
        c.execute('''CREATE TABLE IF NOT EXISTS users (
                        id INTEGER PRIMARY KEY AUTOINCREMENT,
                        [Nome] TEXT,
                        [Email] TEXT,
                        [Account_Type] TEXT,
                        [Password] TEXT  
                    )''')
        c.execute('''CREATE TABLE IF NOT EXISTS pagamentos_eventos (
                        id INTEGER PRIMARY KEY AUTOINCREMENT,
                        [Evento_id] INTEGER,
                        [Participante_id] INTEGER,
                        [Data do Evento] DATE,
                        [Data do Pagamento] DATE,
                        [Tipo Evento] TEXT,
                        [Tipo Pagamento] TEXT,  
                        [Forma de Pagamento] TEXT,
                        [Valor Pago] REAL,
                        [Taxa da Máquina] REAL,
                        [Valor Recebido] REAL,
                        [Observação] TEXT,
                        [Status] TEXT
                    )''')
        c.execute('''CREATE TABLE IF NOT EXISTS caixa (
                        id INTEGER PRIMARY KEY AUTOINCREMENT,
                        [Participante_id] INTEGER,
                        [Data] DATE,
                        [Origem] TEXT,
                        [Observação] TEXT,
                        [Valor] REAL,
                        [Operação] TEXT   
                    )''')
        c.execute('''CREATE TABLE IF NOT EXISTS eventos (
                        id INTEGER PRIMARY KEY AUTOINCREMENT,
                        [Data] DATE,
                        [Data do Evento] DATE,
                        [Nome] TEXT,
                        [Telefone] TEXT,
                        [Email] TEXT,
                        [Endereço] TEXT,	
                        [CPF] TEXT,
                        [Tipo Evento] TEXT,
                        [Tipo Pagamento] TEXT,
                        [Status] TEXT
                    )''')
        c.execute('''CREATE TABLE IF NOT EXISTS participantes (
                        id INTEGER PRIMARY KEY AUTOINCREMENT,
                        [Evento_id] INTEGER,
                        [Data] DATE,
                        [Nome] TEXT,
                        [Tipo] TEXT,
                        [Telefone] TEXT,
                        [Email] TEXT,
                        [Endereço] TEXT,
                        [CPF] TEXT,
                        [Status] TEXT
                    )''')
        c.execute('''CREATE TABLE IF NOT EXISTS orcamentos_meninos (
                        id INTEGER PRIMARY KEY AUTOINCREMENT,
                        [Evento_id] INTEGER,    
                        [Participante_id] INTEGER,
                        [Data] DATE,
                        [Ombro-Punho] FLOAT,
                        [Bainha-Calça] FLOAT,
                        [Modelo] TEXT,
                        [Acessórios] TEXT,
                        [Observação] TEXT, 
                        [Valor Total] REAL,
                        [Taxa de Desconto] REAL,
                        [Valor com Desconto] REAL,
                        [Data Retirada] DATE,
                        [Estado Retirada] TEXT,
                        [Contrato Retirada] TEXT,
                        [Status Contrato Retirada] TEXT,
                        [Data Devolução] DATE,
                        [Estado Devolução] TEXT,
                        [Contrato Devolução] TEXT,
                        [Status Contrato Devolução] TEXT,
                        [Status] TEXT	
                    )''')
        c.execute('''CREATE TABLE IF NOT EXISTS orcamentos_meninas (
                        id INTEGER PRIMARY KEY AUTOINCREMENT,
                        [Evento_id] INTEGER,    
                        [Participante_id] INTEGER,
                        [Data] DATE,
                        [Busto] FLOAT,
                        [Cintura] FLOAT,
                        [Ombro-Cintura] FLOAT, 
                        [Cintura-Pé] FLOAT,
                        [Modelo] TEXT,
                        [Acessórios] TEXT,
                        [Observação] TEXT,  
                        [Valor Total] REAL,
                        [Taxa de Desconto] REAL,
                        [Valor com Desconto] REAL,
                        [Data Retirada] DATE,
                        [Estado Retirada] TEXT,
                        [Contrato Retirada] TEXT,
                        [Status Contrato Retirada] TEXT,
                        [Data Devolução] DATE,
                        [Estado Devolução] TEXT,
                        [Contrato Devolução] TEXT,
                        [Status Contrato Devolução] TEXT,
                        [Status] TEXT
                    )''')
        # Insert admin user only if the users table is empty
        c.execute("SELECT COUNT(*) FROM users")
        if c.fetchone()[0] == 0:
            c.execute("INSERT INTO users (Nome, Email, Account_Type, Password) VALUES (?, ?, ?, ?)", ("Admin", "admin@micas.com.br", "Admin", hash_password("123456")))
        conn.commit()        

def ensure_month_year_folder(base_path, date):
    """
    Ensures a folder named 'month_year' exists in the given path. If not, creates it.

    Args:
        base_path (str): The base directory path.
        date (datetime.date): The date to use for the folder name.

    Returns:
        str: The full path to the 'month_year' folder.
    """
    # Format the folder name as "month_year" (e.g., "April_2025")
    folder_name = date.strftime("%m_%Y")
    folder_path = os.path.join(base_path, folder_name)

    # Check if the folder exists, and create it if not
    if not os.path.exists(folder_path):
        os.makedirs(folder_path)

    return folder_path

# Função para exportar os dados mantendo formatação
def exportar_para_excel(df, file_path, min_date=datetime.datetime.today().date(), max_date=datetime.datetime.today().date()):
    wb = Workbook()
    ws = wb.active

    # Add "Caixa Enviado" in bold in the first column and max date in the second column of the first row
    bold_font = Font(bold=True)
    ws['A1'] = "Caixa Enviado"
    ws['A1'].font = bold_font
    ws['B1'] = min_date.strftime("%Y-%m-%d")
    ws['B1'].font = bold_font
    ws['C1'] = max_date.strftime("%Y-%m-%d")
    ws['C1'].font = bold_font

    green_fill = PatternFill(start_color="A8D08D", end_color="A8D08D", fill_type="solid")
    thin_border = Border(left=Side(style='thin'), 
                         right=Side(style='thin'), 
                         top=Side(style='thin'), 
                         bottom=Side(style='thin'))
    thick_border = Border(left=Side(style='thick'), 
                          right=Side(style='thick'), 
                          top=Side(style='thick'), 
                          bottom=Side(style='thick'))
    
    # Add column headers starting from row 4
    ws.append([""] * len(df.columns))  # Empty row 2
    ws.append([""] * len(df.columns))  # Empty row 3
    ws.append(df.columns.tolist())
    for cell in ws[ws.max_row]:
        cell.border = thick_border

    # Adjust column widths based on column names
    for col_idx, column_name in enumerate(df.columns, start=1):
        column_letter = ws.cell(row=4, column=col_idx).column_letter
        ws.column_dimensions[column_letter].width = max(len(column_name) + 2, 15)  # Add padding for better visibility

    # Add data rows
    for row in df.itertuples(index=False):
        ws.append(row)
        for cell in ws[ws.max_row]:
            cell.border = thin_border
            if row[4] == "Depósito":  # Índice da coluna "Forma de Pagamento"
                cell.fill = green_fill

    wb.save(file_path)

def is_valid_email(email):
    # Regex to validate email format
    return re.match(r'^[a-z0-9._%+-]+@[a-z0-9.-]+\.[a-z]{2,}$', email, re.IGNORECASE)

def is_valid_password(password):
    # Password must be at least 8 characters, contain one lowercase, one uppercase, and one digit
    return len(password) >= 8 and \
           any(c.islower() for c in password) and \
           any(c.isupper() for c in password) and \
           any(c.isdigit() for c in password)   

def is_valid_cpf(cpf):
    """
    Validates CPF format: xxx.xxx.xxx-xx and ensures it has 11 digits.
    """
    # Remove non-numeric characters
    cpf_digits = re.sub(r'\D', '', cpf)
    # Check if it has exactly 11 digits
    return len(cpf_digits) == 11 and cpf_digits.isdigit()

def is_valid_telefone(telefone):
    """
    Validates phone number format: (xx) xxxxx-xxxx and ensures it has 11 digits.
    """
    # Remove non-numeric characters
    telefone_digits = re.sub(r'\D', '', telefone)
    # Check if it has exactly 11 digits
    return len(telefone_digits) == 11 and telefone_digits.isdigit()

def format_telefone(telefone):
    """
    Formats a valid 11-digit phone number into (xx) xxxxx-xxxx.
    """
    telefone_digits = re.sub(r'\D', '', telefone)
    return f"({telefone_digits[:2]}) {telefone_digits[2:7]}-{telefone_digits[7:]}" 

